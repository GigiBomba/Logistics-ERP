from __future__ import annotations

import contextlib
import csv
import logging
import os
import threading
import warnings
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, mm
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from config import Config
from models.common import ErrorDetail, ServiceResult
from models.export_models import ExportOperationResult, ExportRequest, ExportResult
from utils.helpers import remove_accents

logger = logging.getLogger(__name__)

_ALLOWED_EXTENSIONS = {".pdf", ".xlsx", ".csv"}


class ExportService:
    """Export service with typed Pydantic model support.

    New usage (preferred):
        service = ExportService(db=db)
        result = service.export(
            ExportRequest(format="pdf", entity_type="trip", ...), user_id=42
        )

    Old usage (deprecated):
        service = ExportService()
        path = service.generate_pdf(trips, filename)
        path = service.generate_excel(trips, filename)
    """

    def __init__(self, prefs=None, db=None):
        self.styles = getSampleStyleSheet()
        self.reports_dir = Config.REPORTS_DIR
        if not os.path.exists(self.reports_dir):
            os.makedirs(self.reports_dir)
        self.prefs = prefs
        self.db = db

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _safe_filename(filename: str, allowed_ext: str = ".pdf") -> str:
        """Validate filename is a plain name with allowed extension (prevents path traversal)."""
        name = os.path.basename(filename)
        ext = os.path.splitext(name)[1].lower()
        if ext not in _ALLOWED_EXTENSIONS and ext != allowed_ext:
            raise ValueError(f"Filename extension '{ext}' not allowed")
        if ".." in name or "/" in name or "\\" in name:
            raise ValueError(f"Invalid filename: {filename!r}")
        return name

    def _ensure_reports_dir(self) -> None:
        """Create reports directory if it does not exist."""
        if not os.path.exists(self.reports_dir):
            os.makedirs(self.reports_dir)

    # ------------------------------------------------------------------
    # Old methods — kept for backward compatibility with deprecation warning
    # ------------------------------------------------------------------

    def generate_pdf(self, trips_or_request, filename_or_user_id=None, **kwargs):
        """Generate PDF export.

        New usage (typed):
            generate_pdf(request: ExportRequest, user_id: int) -> ExportOperationResult
        Old usage (deprecated):
            generate_pdf(trips, filename=None) -> str
        """
        # Support old ``filename`` keyword argument
        if "filename" in kwargs and filename_or_user_id is None:
            filename_or_user_id = kwargs.pop("filename")

        # New typed path
        if isinstance(trips_or_request, ExportRequest):
            if filename_or_user_id is None:
                raise TypeError("user_id is required when calling generate_pdf with ExportRequest")
            return self.export(trips_or_request, filename_or_user_id)

        # Old deprecated path
        warnings.warn(
            "generate_pdf(trips, filename) is deprecated \u2014 use export(ExportRequest, user_id) instead",
            DeprecationWarning,
            stacklevel=2,
        )
        logger.info(
            "Deprecated generate_pdf called with %d trips",
            len(trips_or_request) if trips_or_request else 0,
        )

        trips = trips_or_request
        if not filename_or_user_id:
            filename = f"raport_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        else:
            filename = self._safe_filename(filename_or_user_id, allowed_ext=".pdf")

        full_path = os.path.join(self.reports_dir, filename)

        doc = SimpleDocTemplate(
            full_path,
            pagesize=A4,
            leftMargin=1 * cm,
            rightMargin=1 * cm,
            topMargin=1 * cm,
            bottomMargin=1 * cm,
        )
        story = []

        title_style = ParagraphStyle(
            "T", parent=self.styles["Title"], fontSize=18, textColor=colors.HexColor("#1a73e8")
        )
        story.append(Paragraph(remove_accents("Cashflow Manager - Raport Activitate"), title_style))
        story.append(HRFlowable(width="100%", color=colors.HexColor("#1a73e8"), spaceAfter=20))

        data = [["Data", "Camion", "Sofer", "Client", "KM", "Brut/km", "Profit", "Status"]]

        for t in trips:
            data.append([
                str(t["created_at"])[:10],
                remove_accents(t["truck_number"]),
                remove_accents(t["driver_name"]),
                remove_accents(t["client_name"]),
                f"{(t.get('distance_km') or 0):.0f}",
                f"{(t.get('gross_per_km') or 0):.2f}",
                f"{(t.get('net_profit') or 0):.2f}",
                remove_accents(t["status"]),
            ])

        table = Table(
            data,
            colWidths=[2.2 * cm, 2.5 * cm, 3.2 * cm, 4 * cm, 1.5 * cm, 1.8 * cm, 2.3 * cm, 2.3 * cm],
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(table)

        try:
            doc.build(story)
        except (OSError, ValueError, RuntimeError):
            logger.exception("PDF build failed")
            if os.path.exists(full_path):
                with contextlib.suppress(OSError):
                    os.remove(full_path)
            raise

        logger.info("Deprecated generate_pdf wrote file: %s", full_path)
        return full_path

    def generate_excel(self, trips_or_request, filename_or_user_id=None, **kwargs):
        """Generate Excel export.

        New usage (typed):
            generate_excel(request: ExportRequest, user_id: int) -> ExportOperationResult
        Old usage (deprecated):
            generate_excel(trips, filename=None) -> str
        """
        # Support old ``filename`` keyword argument
        if "filename" in kwargs and filename_or_user_id is None:
            filename_or_user_id = kwargs.pop("filename")

        # New typed path
        if isinstance(trips_or_request, ExportRequest):
            if filename_or_user_id is None:
                raise TypeError("user_id is required when calling generate_excel with ExportRequest")
            req = trips_or_request
            return self.export(
                ExportRequest(
                    format="excel",
                    entity_type=req.entity_type,
                    entity_id=req.entity_id,
                    entity_ids=req.entity_ids,
                    template=req.template,
                    filename=req.filename,
                    include_logo=req.include_logo,
                    language=req.language,
                ),
                filename_or_user_id,
            )

        # Old deprecated path
        warnings.warn(
            "generate_excel(trips, filename) is deprecated \u2014 use export(ExportRequest, user_id) instead",
            DeprecationWarning,
            stacklevel=2,
        )
        logger.info(
            "Deprecated generate_excel called with %d trips",
            len(trips_or_request) if trips_or_request else 0,
        )

        trips = trips_or_request
        if not filename_or_user_id:
            filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        else:
            filename = self._safe_filename(filename_or_user_id, allowed_ext=".xlsx")

        full_path = os.path.join(self.reports_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Istoric Curse"

        headers = [
            "ID", "Data", "Camion", "Sofer", "Client", "KM",
            "Pret Total", "Profit Net", "Brut/km", "Net/km",
            "Status", "Combustibil", "Taxe", "Salariu",
        ]
        ws.append(headers)

        for t in trips:
            ws.append([
                t["id"],
                t["created_at"],
                t["truck_number"],
                t["driver_name"],
                t["client_name"],
                t["distance_km"],
                t["total_price_eur"],
                t["net_profit"],
                t["gross_per_km"],
                t["rate_per_km"],
                t["status"],
                t["fuel_cost"],
                t["toll_cost"],
                t["salary_cost"],
            ])

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1A73E8", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                with contextlib.suppress(Exception):
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        wb.save(full_path)
        logger.info("Deprecated generate_excel wrote file: %s", full_path)
        return full_path

    # ------------------------------------------------------------------
    # New typed methods
    # ------------------------------------------------------------------

    def export(self, request: ExportRequest, user_id: int) -> ExportOperationResult:
        """Main typed export entry point.

        Checks permissions, routes to the correct generator based on
        ``request.format``, and returns a ``ServiceResult[ExportResult]``.

        Args:
            request: Typed export parameters.
            user_id: ID of the user requesting the export.

        Returns:
            ExportOperationResult with file metadata on success,
            or error detail on failure.
        """
        logger.info(
            "Export requested \u2014 format=%s entity_type=%s entity_ids=%s user_id=%d",
            request.format,
            request.entity_type,
            request.entity_ids,
            user_id,
        )

        # Permission check
        perm_check = self._check_export_permission(user_id)
        if not perm_check.success:
            return perm_check

        # Route to format-specific handler
        try:
            if request.format == "pdf":
                return self._generate_pdf_export(request, user_id)
            elif request.format == "excel":
                return self._generate_excel_export(request, user_id)
            elif request.format == "csv":
                return self._generate_csv_export(request, user_id)
            else:
                return ServiceResult(
                    success=False,
                    errors=[
                        ErrorDetail(
                            message=f"Unsupported export format: {request.format}",
                            code="INVALID_FORMAT",
                        )
                    ],
                )
        except (ValueError, TypeError, RuntimeError) as exc:
            logger.error("Export failed for user %d: %s", user_id, exc, exc_info=True)
            return ServiceResult(
                success=False,
                errors=[ErrorDetail(message=str(exc), code="EXPORT_FAILED")],
            )

    def generate_csv(self, request: ExportRequest, user_id: int) -> ExportOperationResult:
        """Dedicated CSV export.

        Shortcut that sets ``format="csv"`` on the request and delegates
        to :meth:`export`.
        """
        logger.info("CSV export requested \u2014 entity_type=%s user_id=%d", request.entity_type, user_id)
        csv_request = ExportRequest(
            format="csv",
            entity_type=request.entity_type,
            entity_id=request.entity_id,
            entity_ids=request.entity_ids,
            template=request.template,
            filename=request.filename or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            include_logo=request.include_logo,
            language=request.language,
        )
        return self.export(csv_request, user_id)

    def generate_cmr_pdf(self, trip_id: int, user_id: int) -> ExportOperationResult:
        """Convenience method for CMR PDF export of a single trip.

        Builds an ``ExportRequest`` targeting the CMR template and
        delegates to :meth:`export`.
        """
        logger.info("CMR PDF export requested \u2014 trip_id=%d user_id=%d", trip_id, user_id)
        request = ExportRequest(
            format="pdf",
            entity_type="cmr",
            entity_ids=[trip_id],
            template="cmr",
            filename=f"cmr_trip_{trip_id}.pdf",
        )
        return self.export(request, user_id)

    # ------------------------------------------------------------------
    # Async execution
    # ------------------------------------------------------------------

    def generate_pdf_async(
        self,
        request: ExportRequest,
        user_id: int,
        callback,
    ) -> threading.Thread:
        """Generate PDF export in a background thread.

        Args:
            request: Typed export parameters.
            user_id: ID of the user requesting the export.
            callback: Callable that receives the ``ExportOperationResult``
                      when generation completes.

        Returns:
            The background ``threading.Thread`` (daemon) for optional join.
        """
        def _run():
            try:
                result = self.export(request, user_id)
                callback(result)
            except (ValueError, TypeError, RuntimeError) as e:
                logger.error("Async PDF export failed: %s", e, exc_info=True)
                callback(ServiceResult(
                    success=False,
                    errors=[ErrorDetail(message=str(e), code="ASYNC_ERROR")],
                ))

        thread = threading.Thread(target=_run, daemon=True)
        thread.start()
        return thread

    # ------------------------------------------------------------------
    # Permission helper
    # ------------------------------------------------------------------

    def _check_export_permission(self, user_id: int) -> ExportOperationResult:
        """Check whether *user_id* has the ``can_export_data`` permission.

        Returns an error ``ServiceResult`` if denied, or a successful one
        (with ``data=None``) if allowed.
        """
        try:
            from services.permission_service import PermissionService

            perm = PermissionService(self.db)
            check = perm.can_export_data(user_id)
            if not check.allowed:
                logger.error("Export permission denied for user %d: %s", user_id, check.reason)
                return ServiceResult(
                    success=False,
                    errors=[ErrorDetail(message=check.reason, code="PERMISSION_DENIED")],
                )
            return ServiceResult(success=True)
        except (ValueError, RuntimeError, TypeError) as exc:
            logger.error("Permission check failed for user %d: %s", user_id, exc, exc_info=True)
            return ServiceResult(
                success=False,
                errors=[ErrorDetail(message=f"Permission check error: {exc}", code="PERMISSION_ERROR")],
            )

    def save_binary(self, path: str, data: bytes) -> None:
        """Write arbitrary binary data to a file.

        Used for exporting route files (``.operionroute``) where the
        serialised payload is produced by another service.
        """
        with open(path, "wb") as f:
            f.write(data)

    # ------------------------------------------------------------------
    # Internal export implementations
    # ------------------------------------------------------------------

    def _generate_pdf_export(self, request: ExportRequest, user_id: int) -> ExportOperationResult:
        """Internal PDF export implementation."""
        entities = self._fetch_entities(request)
        now = datetime.now()

        if not request.filename:
            filename = f"export_{now.strftime('%Y%m%d_%H%M%S')}.pdf"
        else:
            filename = self._safe_filename(request.filename, allowed_ext=".pdf")

        self._ensure_reports_dir()
        full_path = os.path.join(self.reports_dir, filename)

        doc = SimpleDocTemplate(
            full_path,
            pagesize=A4,
            leftMargin=1 * cm,
            rightMargin=1 * cm,
            topMargin=1 * cm,
            bottomMargin=1 * cm,
        )
        story = []

        title_style = ParagraphStyle(
            "T", parent=self.styles["Title"], fontSize=18, textColor=colors.HexColor("#1a73e8")
        )
        title_text = f"Export {request.entity_type.upper()} - {now.strftime('%d.%m.%Y')}"
        story.append(Paragraph(remove_accents(title_text), title_style))
        story.append(HRFlowable(width="100%", color=colors.HexColor("#1a73e8"), spaceAfter=20))

        if entities and isinstance(entities[0], dict):
            headers = [remove_accents(str(k)) for k in entities[0].keys()]
            data = [headers]
            for row in entities:
                data.append([
                    remove_accents(str(v)) if isinstance(v, str) else str(v) for v in row.values()
                ])
        else:
            data = [["No data"]]

        col_width = (19 * cm) / max(len(data[0]), 1)
        col_widths = [min(col_width, 4 * cm)] * len(data[0])

        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(table)

        try:
            doc.build(story)
        except (OSError, ValueError, RuntimeError):
            logger.exception("PDF export build failed")
            if os.path.exists(full_path):
                with contextlib.suppress(OSError):
                    os.remove(full_path)
            raise

        file_size = os.path.getsize(full_path)
        logger.info("PDF export completed \u2014 path=%s size=%d bytes", full_path, file_size)

        return ServiceResult(
            success=True,
            data=ExportResult(
                file_path=full_path,
                format="pdf",
                entity_type=request.entity_type,
                file_size=file_size,
                generated_at=now,
            ),
        )

    def _generate_excel_export(self, request: ExportRequest, user_id: int) -> ExportOperationResult:
        """Internal Excel export implementation."""
        entities = self._fetch_entities(request)
        now = datetime.now()

        if not request.filename:
            filename = f"export_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = self._safe_filename(request.filename, allowed_ext=".xlsx")

        self._ensure_reports_dir()
        full_path = os.path.join(self.reports_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Export {request.entity_type[:20]}"

        if entities and isinstance(entities[0], dict):
            headers = list(entities[0].keys())
            ws.append(headers)
            for row in entities:
                ws.append([row.get(h, "") for h in headers])
        else:
            ws.append(["No data"])

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1A73E8", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                with contextlib.suppress(TypeError, ValueError, AttributeError):
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(full_path)

        file_size = os.path.getsize(full_path)
        logger.info("Excel export completed \u2014 path=%s size=%d bytes", full_path, file_size)

        return ServiceResult(
            success=True,
            data=ExportResult(
                file_path=full_path,
                format="excel",
                entity_type=request.entity_type,
                file_size=file_size,
                generated_at=now,
            ),
        )

    def _generate_csv_export(self, request: ExportRequest, user_id: int) -> ExportOperationResult:
        """Internal CSV export implementation."""
        entities = self._fetch_entities(request)
        now = datetime.now()

        if not request.filename:
            filename = f"export_{now.strftime('%Y%m%d_%H%M%S')}.csv"
        else:
            filename = self._safe_filename(request.filename, allowed_ext=".csv")

        self._ensure_reports_dir()
        full_path = os.path.join(self.reports_dir, filename)

        with open(full_path, "w", newline="", encoding="utf-8-sig") as f:
            if entities and isinstance(entities[0], dict):
                writer = csv.DictWriter(f, fieldnames=list(entities[0].keys()))
                writer.writeheader()
                writer.writerows(entities)
            else:
                writer = csv.writer(f)
                writer.writerow(["No data"])

        file_size = os.path.getsize(full_path)
        logger.info("CSV export completed \u2014 path=%s size=%d bytes", full_path, file_size)

        return ServiceResult(
            success=True,
            data=ExportResult(
                file_path=full_path,
                format="csv",
                entity_type=request.entity_type,
                file_size=file_size,
                generated_at=now,
            ),
        )

    # ------------------------------------------------------------------
    # Dispatch board export helpers
    # ------------------------------------------------------------------

    def generate_dispatch_board_csv(
        self,
        card_data: list[dict[str, Any]],
        file_path: str,
    ) -> str:
        """Write dispatch board *card_data* to *file_path* as CSV.

        Accepts the raw card-data dicts from the board view and writes a
        structured CSV with columns: Trip ID, Status, Truck, Driver, Origin,
        Destination, Departure, ETA, Alerts.

        Returns:
            *file_path* on success (caller should handle exceptions).
        """
        import csv

        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "Trip ID", "Status", "Truck", "Driver", "Origin", "Destination",
                "Departure", "ETA", "Alerts",
            ])
            for cd in card_data:
                writer.writerow([
                    cd.get("trip_id", ""),
                    cd.get("status", ""),
                    cd.get("truck_plate", ""),
                    cd.get("driver_name", ""),
                    cd.get("origin", ""),
                    cd.get("destination", ""),
                    cd.get("departure_date", ""),
                    cd.get("eta", ""),
                    cd.get("alerts_count", 0),
                ])
        return file_path

    def generate_dispatch_board_pdf(
        self,
        card_data: list[dict[str, Any]],
        file_path: str,
    ) -> str:
        """Generate a dark-themed PDF for dispatch board *card_data*.

        Groups trips by status column (Planned / Loading / In Transit /
        Delivered / Cancelled), renders up to 50 trips per group in
        colour-coded tables, and writes the result to *file_path*.

        Returns:
            *file_path* on success (caller should handle exceptions).
        """
        from datetime import datetime

        # Status → column-key mapping (mirrors board_state.STATUS_TO_COLUMN)
        STATUS_TO_COLUMN: dict[str, str] = {
            "Planned": "Planned", "Scheduled": "Planned", "Pending": "Planned",
            "Loading": "Loading", "Preparing": "Loading", "Pickup": "Loading",
            "In Transit": "In Transit", "InTransit": "In Transit",
            "Active": "In Transit", "InProgress": "In Transit",
            "Delivered": "Delivered", "Completed": "Delivered", "Done": "Delivered",
            "Invoiced": "Delivered", "Paid": "Delivered",
            "Cancelled": "Cancelled",
        }

        status_colors: dict[str, Any] = {
            "Planned": colors.HexColor("#1c1917"),
            "Loading": colors.HexColor("#341a00"),
            "In Transit": colors.HexColor("#0f1f4a"),
            "Delivered": colors.HexColor("#052e16"),
            "Cancelled": colors.HexColor("#1A1A20"),
        }

        header_style = ParagraphStyle(
            "Header", textColor=colors.HexColor("#fafafa"),
            fontSize=9, fontName="Helvetica-Bold",
        )
        cell_style = ParagraphStyle(
            "Cell", textColor=colors.HexColor("#a1a1aa"), fontSize=8,
        )

        doc = SimpleDocTemplate(
            file_path, pagesize=landscape(A4),
            topMargin=10 * mm, bottomMargin=10 * mm,
        )
        styles = getSampleStyleSheet()
        elements: list = []

        title_style = ParagraphStyle(
            "Title", parent=styles["Title"], fontSize=14,
            textColor=colors.HexColor("#fafafa"),
        )
        elements.append(
            Paragraph(
                f"Dispatch Board \u2014 {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                title_style,
            )
        )
        elements.append(Spacer(1, 6 * mm))

        for col_key in ["Planned", "Loading", "In Transit", "Delivered", "Cancelled"]:
            col_trips = [
                cd for cd in card_data
                if STATUS_TO_COLUMN.get(cd.get("status", "")) == col_key
            ]
            bg = status_colors.get(col_key, colors.grey)

            elements.append(Paragraph(
                f"{col_key} ({len(col_trips)})", header_style,
            ))
            elements.append(Spacer(1, 2 * mm))

            if col_trips:
                table_data = [
                    ["Trip ID", "Truck", "Driver", "Route", "Departure", "ETA"],
                ]
                for cd in col_trips[:50]:
                    table_data.append([
                        cd.get("trip_id", ""),
                        cd.get("truck_plate", ""),
                        cd.get("driver_name", ""),
                        f"{cd.get('origin', '?')} \u2192 {cd.get('destination', '?')}",
                        cd.get("departure_date", ""),
                        cd.get("eta", ""),
                    ])
                tbl = Table(
                    table_data,
                    colWidths=[45 * mm, 40 * mm, 45 * mm, 60 * mm, 40 * mm, 40 * mm],
                )
                tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), bg),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#fafafa")),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#27272a")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [colors.HexColor("#111113"), colors.HexColor("#18181b")]),
                ]))
                elements.append(tbl)
            else:
                elements.append(Paragraph("No trips", cell_style))
            elements.append(Spacer(1, 4 * mm))

        doc.build(elements)
        return file_path

    # ------------------------------------------------------------------
    # Data fetching
    # ------------------------------------------------------------------

    def _fetch_entities(self, request: ExportRequest) -> list[dict[str, Any]]:
        """Fetch entity data from the database based on ``entity_type``.

        Args:
            request: Export request containing entity type and optional IDs.

        Returns:
            List of row dicts.

        Raises:
            ValueError: If the entity type is unsupported or db is not set.
        """
        if not self.db:
            raise ValueError("Database connection (db) is required for entity export")

        et = request.entity_type

        if et in ("trip", "cmr"):
            from repositories.trip_repository import TripRepository

            repo = TripRepository(self.db)
            if request.entity_ids:
                return repo.get_by_ids(request.entity_ids)
            return repo.get_all()

        elif et == "invoice":
            from repositories.invoice_repository import InvoiceRepository

            repo = InvoiceRepository(self.db)
            if request.entity_ids:
                items = []
                for eid in request.entity_ids:
                    row = repo.get_by_id(eid)
                    if row:
                        items.append(row)
                return items
            return repo.get_all()

        elif et == "receipt":
            from repositories.receipt_repository import ReceiptRepository

            repo = ReceiptRepository(self.db)
            if request.entity_ids:
                items = []
                for eid in request.entity_ids:
                    row = repo.get_by_id(eid)
                    if row:
                        items.append(row)
                return items
            return repo.get_all()

        else:
            raise ValueError(f"Unsupported entity_type for export: {et}")
