"""Orchestrates data export (Tab 3) for the Migration Center.

Supports exporting TRIP, CLIENT, DRIVER, TRUCK, and INVOICE entities
to CSV, Excel, and JSON formats.  Reuses ``CsvService`` and the
existing repository layer for data access.
"""

from __future__ import annotations

import json
import logging
import os
from typing import Any

from database.db_manager import DatabaseManager
from services.migration.types import EntityType, ExportFormat, ProgressCallback

logger = logging.getLogger(__name__)


class EmigrateService:
    """Service for exporting data from the system.

    Typical usage::

        svc = EmigrateService(db)
        path = svc.export(
            EntityType.TRIP, ExportFormat.CSV, "/tmp/export.csv",
            filters={"status": "Delivered"},
            field_selection=["id", "client_name", "total_price_eur"],
        )
    """

    SUPPORTED_ENTITIES: list[EntityType] = [
        EntityType.TRIP,
        EntityType.CLIENT,
        EntityType.DRIVER,
        EntityType.TRUCK,
        EntityType.INVOICE,
    ]

    SUPPORTED_FORMATS: list[ExportFormat] = [
        ExportFormat.CSV,
        ExportFormat.EXCEL,
        ExportFormat.JSON,
    ]

    def __init__(self, db: DatabaseManager) -> None:
        self.db = db
        self._repos: dict[EntityType, Any] = {}

    # ── Repo cache ─────────────────────────────────────────────────────

    def _get_repo(self, entity_type: EntityType) -> Any:
        """Lazy-load and cache the repository for *entity_type*."""
        if entity_type not in self._repos:
            if entity_type == EntityType.CLIENT:
                from repositories.client_repository import ClientRepository
                self._repos[entity_type] = ClientRepository(self.db)
            elif entity_type == EntityType.DRIVER:
                from repositories.driver_repository import DriverRepository
                self._repos[entity_type] = DriverRepository(self.db)
            elif entity_type == EntityType.TRUCK:
                from repositories.fleet_repository import FleetRepository
                self._repos[entity_type] = FleetRepository(self.db)
            elif entity_type == EntityType.TRIP:
                from repositories.trip_repository import TripRepository
                self._repos[entity_type] = TripRepository(self.db)
            elif entity_type == EntityType.INVOICE:
                from repositories.invoice_repository import InvoiceRepository
                self._repos[entity_type] = InvoiceRepository(self.db)
            else:
                raise ValueError(f"No repository available for entity type: {entity_type}")
        return self._repos[entity_type]

    # ── Counting records ────────────────────────────────────────────────

    def count_records(self, entity_type: EntityType, filters: dict | None = None) -> int:
        """Return the number of records for *entity_type*, optionally filtered."""
        repo = self._get_repo(entity_type)
        if not repo:
            return 0
        try:
            if hasattr(repo, "count"):
                return repo.count(filters)
            rows = repo._fetchall(
                f"SELECT COUNT(*) as cnt FROM {repo.TABLE} {repo._company_filter()}",
                repo._company_params(),
            )
            return rows[0]["cnt"] if rows else 0
        except Exception:
            return 0

    # ── Main export entry point ────────────────────────────────────────

    def export(
        self,
        entity_type: EntityType,
        fmt: ExportFormat,
        output_path: str,
        filters: dict[str, Any] | None = None,
        field_selection: list[str] | None = None,
        progress_cb: ProgressCallback = None,
    ) -> str:
        """Export entity data to a file.

        Args:
            entity_type: The type of entity to export.
            fmt: Target export format.
            output_path: Absolute or relative destination path.
            filters: Optional dict of column → value to filter rows.
            field_selection: Optional list of column names to include.
            progress_cb: Optional progress callback.

        Returns:
            The absolute path to the written file.

        Raises:
            ValueError: If the entity type or format is not supported.
        """
        if entity_type not in self.SUPPORTED_ENTITIES:
            raise ValueError(
                f"Unsupported entity type for export: {entity_type.value}. "
                f"Supported: {[e.value for e in self.SUPPORTED_ENTITIES]}"
            )
        if fmt not in self.SUPPORTED_FORMATS:
            raise ValueError(
                f"Unsupported export format: {fmt.value}. "
                f"Supported: {[f.value for f in self.SUPPORTED_FORMATS]}"
            )

        if progress_cb:
            progress_cb("fetching", 10, f"Fetching {entity_type.value} data...")

        # ── Fetch data ────────────────────────────────────────────────
        repo = self._get_repo(entity_type)
        try:
            rows = repo.get_all()
        except Exception:
            # Fallback: direct SELECT
            table = getattr(repo, "TABLE", f"{entity_type.value}s")
            try:
                rows = repo._fetchall(f"SELECT * FROM {table}")
            except Exception as exc:
                raise RuntimeError(f"Failed to fetch {entity_type.value} data: {exc}") from exc

        if progress_cb:
            progress_cb("fetching", 40, f"Fetched {len(rows)} rows")

        # ── Apply filters ─────────────────────────────────────────────
        if filters:
            rows = self._apply_filters(rows, filters)

        # ── Apply field selection ─────────────────────────────────────
        if field_selection:
            rows = [
                {k: v for k, v in row.items() if k in field_selection}
                for row in rows
            ]

        if progress_cb:
            progress_cb("writing", 70, f"Writing {len(rows)} rows to {fmt.value}...")

        # ── Write output ──────────────────────────────────────────────
        dirname = os.path.dirname(os.path.abspath(output_path))
        if dirname:
            os.makedirs(dirname, exist_ok=True)

        writers = {
            ExportFormat.CSV: self._write_csv,
            ExportFormat.EXCEL: self._write_excel,
            ExportFormat.JSON: self._write_json,
        }
        writer = writers.get(fmt)
        if writer is None:
            raise ValueError(f"No writer available for format: {fmt.value}")

        try:
            writer(rows, output_path)
        except Exception as exc:
            logger.exception("Export write failed for %s", output_path)
            raise RuntimeError(f"Failed to write export file: {exc}") from exc

        if progress_cb:
            progress_cb("complete", 100, f"Exported {len(rows)} rows to {output_path}")

        abs_path = os.path.abspath(output_path)
        logger.info("Export complete: %s (%d rows, %s)", abs_path, len(rows), fmt.value)
        return abs_path

    # ── Filtering ──────────────────────────────────────────────────────

    @staticmethod
    def _apply_filters(
        rows: list[dict[str, Any]],
        filters: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Filter rows by matching column values (case-insensitive string match)."""
        filtered = list(rows)
        for key, value in filters.items():
            if value is None:
                continue
            str_value = str(value).lower()
            filtered = [
                row
                for row in filtered
                if str(row.get(key, "")).lower() == str_value
            ]
        return filtered

    # ── Format writers ─────────────────────────────────────────────────

    @staticmethod
    def _write_csv(rows: list[dict[str, Any]], path: str) -> None:
        """Write rows to a CSV file using CsvService or the csv module."""
        try:
            from services.csv_service import CsvService

            fieldnames = list(rows[0].keys()) if rows else None
            CsvService.export(rows, path, fieldnames=fieldnames)
        except Exception:
            import csv

            fieldnames = list(rows[0].keys()) if rows else []
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)

    @staticmethod
    def _write_excel(rows: list[dict[str, Any]], path: str) -> None:
        """Write rows to an Excel (.xlsx) file using openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required to export to Excel format")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        if not rows:
            wb.save(path)
            return

        fieldnames = list(rows[0].keys())

        # Header row
        ws.append(fieldnames)
        header_fill = PatternFill(start_color="1A73E8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        for row in rows:
            ws.append([row.get(field, "") for field in fieldnames])

        # Auto-adjust column widths
        for col_idx, field in enumerate(fieldnames, 1):
            max_len = len(field)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True):
                if row[0] is not None:
                    max_len = max(max_len, len(str(row[0])))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 60)

        wb.save(path)

    @staticmethod
    def _write_json(rows: list[dict[str, Any]], path: str) -> None:
        """Write rows to a JSON file."""
        with open(path, "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=2, default=str)
